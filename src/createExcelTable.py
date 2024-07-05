
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def save_in_excel(game_name: str, data: list):
    list_media = ['Twitch', 'Sub Twitch', 'Youtube', 'Sub Youtube', 'Vk', 'Sub Vk', 'Telegram', 'Sub Telegram', 'tiktok', 'Sub Tiktok']
    workbook = Workbook()
    ws = workbook.active
    
    for i in range(1, len(list_media)):
        colm = get_column_letter(i)
        if i % 2:
            ws.column_dimensions[colm].width = 40
        else:
            ws.column_dimensions[colm].width = 12


    ws.append(list_media)

    for social in data:
        ws.append(social)


    try:
        workbook.save(f'{game_name}.xlsx')
    except Exception:
        for i in range(1, 11):
            try:
                workbook.save(f'{game_name}({i}).xlsx')
            except Exception:
                pass
            else:
                break